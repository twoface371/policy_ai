"""
sheet.py — 정해진 양식(엑셀/CSV)으로 법령 목록 받기
================================================================================
자유 문서에서 정규식으로 인용을 뽑는 경로(analyzer)와 나란히 두는 두 번째
입력 경로다. 사용자가 양식에 직접 적으면 추출 단계가 통째로 사라진다 —
법령명 경계를 추측할 일도, 줄바꿈에 이름이 잘릴 일도 없다.

읽은 결과는 analyzer.find_citations와 **같은 모양**으로 돌려준다. 그래야
checker.resolve_citations 이후(조 대조·저장·다운로드)가 그대로 재사용된다.

양식:
    법령명 | 조 | 항 | 비고
    개인정보 보호법 | 제15조 | 1 |
    지능정보화 기본법 | 46조의2 | | 접근성 관련

조·항은 사람이 적는 대로 관대하게 읽는다('15', '제15조', '15조의2' 모두 허용).
읽지 못한 행은 버리지 않고 오류 목록으로 돌려준다 — 조용히 빠지면 사용자는
그 법령을 검사했다고 믿는다.
================================================================================
"""
import csv
import io
import re
from typing import List, Dict, Tuple

# 양식의 열 이름. 순서가 바뀌어도, 열이 더 있어도 이름으로 찾는다.
COL_LAW = "법령명"
COL_ART = "조"
COL_PARA = "항"
COL_NOTE = "비고"
HEADERS = [COL_LAW, COL_ART, COL_PARA, COL_NOTE]

SHEET_DATA = "법령목록"
SHEET_HELP = "작성 예시"

# 비고는 판정에 쓰이지 않는다. 검사 결과를 받았을 때 '내 문서의 어디를
# 고쳐야 하는지' 찾기 위한 메모다. 자유 문서 경로는 원문 주변 40자가 문맥으로
# 자동으로 붙지만 양식 경로에는 그것이 없어서, 그 자리를 사람이 채우는 칸이다.
EXAMPLES = [
    ["개인정보 보호법", "제15조", "1", "본문 제3조제2항"],
    ["지능정보화 기본법", "46조의2", "", "가지번호는 '46조의2'처럼 적습니다"],
    ["국가재정법 시행령", "10", "", "'제'와 '조'는 없어도 됩니다"],
    ["공공기관의 정보공개에 관한 법률", "제9조", "2", "별표 1 / 담당 ○○과"],
]

# 데이터 시트 첫 줄에 미리 채워 두는 예시. 지우지 않고 그대로 제출하면 있지도
# 않은 인용이 결과에 섞이므로, 이 표시가 비고에 남아 있는 행은 파서가 건너뛴다.
# 사용자가 그 줄을 고쳐 쓰면(= 표시를 지우면) 보통 행으로 읽힌다.
EXAMPLE_MARK = "← 예시입니다. 지우거나 고쳐 쓰세요"
EXAMPLE_ROW = ["개인정보 보호법", "제15조", "1", EXAMPLE_MARK]

# '제15조' '15조' '15' '제5조의2' '5의2' '5-2' 를 모두 받는다
_ART = re.compile(r"^\s*제?\s*(\d+)\s*조?\s*(?:의|[-–])?\s*(\d+)?\s*$")
_PARA = re.compile(r"^\s*제?\s*(\d+)\s*항?\s*$")


def parse_article(raw) -> Tuple[int, int]:
    """조 표기 → (조번호, 가지번호). 못 읽으면 (0, 0)."""
    m = _ART.match(str(raw or ""))
    if not m:
        return 0, 0
    return int(m.group(1)), int(m.group(2) or 0)


def parse_para(raw) -> int:
    """항 표기 → 항번호. 비었거나 못 읽으면 0. 판정에는 쓰지 않고 표시만 한다."""
    m = _PARA.match(str(raw or ""))
    return int(m.group(1)) if m else 0


def _cell(v) -> str:
    """엑셀 셀 → 문자열. 숫자로 들어온 조 번호가 '15.0'이 되지 않게 한다."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _rows_from_xlsx(data: bytes) -> List[List[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb[SHEET_DATA] if SHEET_DATA in wb.sheetnames else wb.worksheets[0]
    return [[_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]


def _rows_from_csv(data: bytes) -> List[List[str]]:
    for enc in ("utf-8-sig", "cp949", "utf-8"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="replace")
    return [[c.strip() for c in row] for row in csv.reader(io.StringIO(text))]


def read_rows(data: bytes, ext: str) -> List[List[str]]:
    return _rows_from_csv(data) if ext == ".csv" else _rows_from_xlsx(data)


def parse(data: bytes, ext: str) -> Tuple[List[Dict], List[Dict]]:
    """양식 파일 → (인용 목록, 오류 목록).

    인용은 analyzer.find_citations와 같은 키를 갖는다. 법령후보를 한 개만
    두는 것은 사용자가 이름을 직접 적었기 때문이다 — 경계를 추측할 필요가
    없으니 후보를 늘려 조회를 낭비할 이유가 없다.
    """
    rows = read_rows(data, ext)
    if not rows:
        return [], [{"행": 0, "사유": "빈 파일입니다", "원문": ""}]

    # 머리글 행 찾기 — 앞쪽에 안내 문구가 몇 줄 있어도 견딘다
    head_at, cols = -1, {}
    for i, row in enumerate(rows[:10]):
        names = {v: j for j, v in enumerate(row) if v}
        if COL_LAW in names and COL_ART in names:
            head_at, cols = i, names
            break
    if head_at < 0:
        return [], [{"행": 1,
                     "사유": f"머리글을 찾지 못했습니다 "
                             f"('{COL_LAW}'·'{COL_ART}' 열이 있어야 합니다)",
                     "원문": " | ".join(rows[0][:6])}]

    def get(row, name):
        j = cols.get(name, -1)
        return row[j] if 0 <= j < len(row) else ""

    cites: List[Dict] = []
    errors: List[Dict] = []
    for i, row in enumerate(rows[head_at + 1:], start=head_at + 2):
        law = get(row, COL_LAW)
        art_raw = get(row, COL_ART)
        note = get(row, COL_NOTE)
        if not law and not art_raw:
            continue                     # 빈 줄은 오류가 아니다
        if EXAMPLE_MARK in note:
            continue                     # 지우지 않은 예시 줄
        raw = " | ".join(x for x in (law, art_raw, get(row, COL_PARA)) if x)
        if not law:
            errors.append({"행": i, "사유": "법령명이 비어 있습니다", "원문": raw})
            continue
        no, branch = parse_article(art_raw)
        if not no:
            errors.append({"행": i,
                           "사유": f"조를 읽지 못했습니다: '{art_raw}'"
                                   if art_raw else "조가 비어 있습니다",
                           "원문": raw})
            continue
        para = parse_para(get(row, COL_PARA))
        label = f"제{no}조" + (f"의{branch}" if branch else "") \
                + (f"제{para}항" if para else "")
        cites.append({"법령": law, "법령후보": [law], "조문": label,
                      "조번호": no, "조가지번호": branch, "비고": note,
                      "문맥": f"{i}행" + (f" · {note}" if note else "")})

    if not cites and not errors:
        # 빈 양식을 그대로 올린 경우. 아무 말도 안 하면 '0건 완료'만 보고
        # 잘 된 것인지 뭔가 잘못된 것인지 구분할 수 없다.
        errors.append({"행": head_at + 2, "사유": "데이터 행이 없습니다 "
                                                 "(머리글 아래에 법령·조를 적어주세요)",
                       "원문": ""})
    return cites, errors


# ============================================================
# 빈 양식 만들기
# ============================================================
def build_template() -> bytes:
    """사용자에게 내려줄 빈 양식.

    예시는 별도 시트에 둔다. 데이터 시트에 두면 지우지 않고 제출했을 때
    있지도 않은 인용이 검사 결과에 섞인다.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DATA
    ws.append(HEADERS)
    ws.append(EXAMPLE_ROW)

    head_fill = PatternFill("solid", fgColor="1F3864")
    for j, _ in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=j)
        c.font = Font(color="FFFFFF", bold=True)
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    # 예시 줄은 한눈에 예시로 보이게 흐리고 기울여 둔다
    for j, _ in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=j).font = Font(color="8A8F98", italic=True)
    for j, w in enumerate((46, 14, 8, 34), start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

    hs = wb.create_sheet(SHEET_HELP)
    guide = [
        ["작성 방법"],
        [""],
        [f"1. '{SHEET_DATA}' 시트에 한 줄에 하나씩 적습니다."],
        ["2. 법령명은 문서에 쓴 그대로 적으면 됩니다. 옛 명칭·약칭도 괜찮습니다"],
        ["   — 현행 명칭을 찾아 '개명 의심'으로 알려 줍니다."],
        ["3. 조는 '제15조' '15조' '15' 중 아무 형식이나 됩니다."],
        ["   가지번호는 '46조의2'처럼 적습니다."],
        ["4. 항은 비워도 됩니다. 판정은 조 단위로 하고, 항은 표시용입니다."],
        ["5. 비고는 검사에 쓰이지 않습니다. 결과를 받았을 때 '내 문서의 어디를"],
        ["   고쳐야 하는지' 찾기 위한 메모 칸입니다(본문 위치, 별표 번호, 담당 부서 등)."],
        ["   적어 두면 내려받는 리포트에도 함께 실립니다."],
        ["6. 읽지 못한 줄은 검사 결과에 '양식 오류'로 함께 보여 줍니다."],
        [f"7. '{SHEET_DATA}' 시트 첫 줄은 예시입니다. 지우거나 고쳐 쓰세요"],
        ["   — 그대로 두면 검사에서 건너뜁니다."],
        [""],
        ["작성 예시"],
        HEADERS,
    ]
    for line in guide:
        hs.append(line)
    for ex in EXAMPLES:
        hs.append(ex)
    # 행 번호를 박아 두면 안내 문구를 한 줄 고칠 때마다 어긋난다
    title_at = len(guide) - 1            # '작성 예시'
    header_at = len(guide)               # 그 아래 머리글
    hs.cell(row=1, column=1).font = Font(bold=True, size=13)
    hs.cell(row=title_at, column=1).font = Font(bold=True)
    for j, _ in enumerate(HEADERS, start=1):
        hs.cell(row=header_at, column=j).font = Font(bold=True)
    for j, w in enumerate((46, 14, 8, 34), start=1):
        hs.column_dimensions[get_column_letter(j)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
